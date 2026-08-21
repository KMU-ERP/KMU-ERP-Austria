from __future__ import annotations

import csv
import io
from datetime import datetime
from pathlib import PurePath

import frappe
from frappe import _
from openpyxl import load_workbook

from kmu_erp_austria.bmd_export.services.validation import validate_account_number


NUMBER_HEADERS = {"kto-nr", "konto", "kontonummer", "account number", "account_number"}
NAME_HEADERS = {"bezeichnung", "kontobezeichnung", "account name", "account_name"}


def _file(file_url: str) -> tuple[str, bytes]:
	name = frappe.db.get_value("File", {"file_url": file_url}, "name")
	if not name:
		raise ValueError(_("Uploaded account file does not exist."))
	doc = frappe.get_doc("File", name)
	if not doc.has_permission("read"):
		raise frappe.PermissionError
	content = doc.get_content(encodings=[])
	return doc.file_name, content if isinstance(content, bytes) else content.encode("utf-8")


def _indexes(headers) -> tuple[int, int]:
	normalized = [str(value or "").strip().casefold() for value in headers]
	try:
		number_index = next(index for index, value in enumerate(normalized) if value in NUMBER_HEADERS)
		name_index = next(index for index, value in enumerate(normalized) if value in NAME_HEADERS)
	except StopIteration as exc:
		raise ValueError(_("Account file needs an account-number and account-name column.")) from exc
	return number_index, name_index


def _xlsx(content: bytes):
	workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
	worksheet = workbook.active
	rows = worksheet.iter_rows(values_only=True)
	headers = next(rows, ())
	number_index, name_index = _indexes(headers)
	return [(row[number_index], row[name_index]) for row in rows if any(value is not None for value in row)]


def _csv(content: bytes):
	text = content.decode("utf-8-sig")
	sample = text[:4096]
	dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
	rows = list(csv.reader(io.StringIO(text), dialect))
	if not rows:
		return []
	number_index, name_index = _indexes(rows[0])
	return [(row[number_index], row[name_index]) for row in rows[1:] if any(value.strip() for value in row)]


def preview_accounts(company: str, file_url: str) -> list[dict]:
	filename, content = _file(file_url)
	extension = PurePath(filename).suffix.lower()
	if extension == ".xlsx":
		raw_rows = _xlsx(content)
	elif extension == ".csv":
		raw_rows = _csv(content)
	else:
		raise ValueError(_("Only XLSX and CSV account files are supported."))
	seen = set()
	result = []
	for row_number, (raw_number, raw_name) in enumerate(raw_rows, start=2):
		try:
			number = validate_account_number(str(raw_number or "").split(".0", 1)[0])
			name = str(raw_name or "").strip()
			if not name:
				raise ValueError(_("Account name is empty."))
			if number in seen:
				raise ValueError(_("Duplicate account number in import file."))
			seen.add(number)
			existing = frappe.db.exists("BMD Account", {"company": company, "account_number": number})
			result.append(
				{
					"row": row_number,
					"account_number": number,
					"account_name": name,
					"status": "Existing" if existing else "New",
					"error": "",
				}
			)
		except Exception as exc:
			result.append(
				{
					"row": row_number,
					"account_number": str(raw_number or ""),
					"account_name": str(raw_name or ""),
					"status": "Invalid",
					"error": str(exc),
				}
			)
	return result


def import_accounts(company: str, file_url: str, source: str) -> dict:
	rows = preview_accounts(company, file_url)
	invalid = [row for row in rows if row["status"] == "Invalid"]
	if invalid:
		raise ValueError(_("Account import contains invalid rows; nothing was written."))
	created = 0
	for row in rows:
		if row["status"] != "New":
			continue
		frappe.get_doc(
			{
				"doctype": "BMD Account",
				"company": company,
				"account_number": row["account_number"],
				"account_name": row["account_name"],
				"active": 1,
				"source": source or PurePath(_file(file_url)[0]).name,
				"import_date": datetime.now(),
			}
		).insert()
		created += 1
	return {"created": created, "skipped": len(rows) - created, "total": len(rows)}
